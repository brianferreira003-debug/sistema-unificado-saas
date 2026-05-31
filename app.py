# ═══════════════════════════════════════════════════
# Sistema Unificado SaaS - v3.0
# Multitenant: cada empresa tiene su propia DB
# Login: Flask-Login con roles (admin/usuario)
# ═══════════════════════════════════════════════════

import os
import json
import io
import re
import secrets
from datetime import date, timedelta, datetime
from contextlib import contextmanager
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    send_file, jsonify, make_response, flash, abort
)
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user,
    login_required, current_user
)
from flask_wtf.csrf import CSRFProtect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.security import generate_password_hash, check_password_hash

# ════════════ CONFIGURACIÓN DESDE .env ════════════

app = Flask(__name__)

# Cargar desde .env si existe, sino variables de entorno
def load_env():
    env_path = os.path.join(os.path.dirname(__file__), '.env')
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                line = line.strip()
                line = line.lstrip('\ufeff')  # BOM
                if line and not line.startswith('#') and '=' in line:
                    key, _, value = line.partition('=')
                    key = key.strip()
                    value = value.strip().strip('"').strip("'")
                    if key and key not in os.environ:
                        os.environ[key] = value

load_env()

app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# ════════════ DIRECTORIOS ════════════

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
TENANT_DIR = os.path.join(DATA_DIR, 'tenants')
CONFIG_FILE = os.path.join(BASE_DIR, 'config.json')

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(TENANT_DIR, exist_ok=True)

# ════════════ LOGIN ════════════

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Iniciá sesión para acceder.'
login_manager.login_message_category = 'warning'

# ════════════ CSRF PROTECTION ════════════

csrf = CSRFProtect(app)

# ════════════ RATE LIMITING ════════════

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://",
)

# ════════════ MODELO USUARIO (DB maestra) ════════════

MASTER_DB = os.path.join(DATA_DIR, 'master.db')

def get_master_db():
    conn = sqlite3.connect(MASTER_DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_master_db():
    conn = get_master_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS empresas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL UNIQUE,
            slug TEXT NOT NULL UNIQUE,
            db_path TEXT NOT NULL,
            activa INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            empresa_id INTEGER NOT NULL REFERENCES empresas(id),
            username TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            nombre TEXT DEFAULT '',
            rol TEXT DEFAULT 'usuario',
            activo INTEGER DEFAULT 1,
            force_password_change INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(empresa_id, username)
        );
    """)
    conn.commit()

    # Agregar columna force_password_change si no existe (migración para DBs existentes)
    cols = [row[1] for row in conn.execute("PRAGMA table_info(usuarios)").fetchall()]
    if 'force_password_change' not in cols:
        conn.execute("ALTER TABLE usuarios ADD COLUMN force_password_change INTEGER DEFAULT 1")
        conn.commit()

    # Crear admin global si no existe
    admin = conn.execute("SELECT id FROM usuarios WHERE rol='admin'").fetchone()
    if not admin:
        # Empresa default para admin
        slug = 'admin'
        db_path = os.path.join(TENANT_DIR, f'{slug}.db')
        conn.execute(
            "INSERT INTO empresas (nombre, slug, db_path) VALUES (?,?,?)",
            ('Administración', slug, db_path)
        )
        empresa_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Admin con hash seguro (cambiar en producción!)
        pw_hash = generate_password_hash('admin123')
        conn.execute(
            "INSERT INTO usuarios (empresa_id, username, password_hash, nombre, rol, force_password_change) VALUES (?,?,?,?,?,?)",
            (empresa_id, 'admin', pw_hash, 'Administrador', 'admin', 1)
        )
        conn.commit()
        print("\n  Admin creado: admin / admin123  (CAMBIAR EN PRODUCCIÓN - force_password_change activo)\n")

    conn.close()

import sqlite3

init_master_db()

class User(UserMixin):
    def __init__(self, id, empresa_id, username, nombre, rol, force_password_change=0):
        self.id = id
        self.empresa_id = empresa_id
        self.username = username
        self.nombre = nombre
        self.rol = rol
        self.force_password_change = force_password_change

    @property
    def is_admin(self):
        return self.rol == 'admin'

@login_manager.user_loader
def load_user(user_id):
    conn = get_master_db()
    row = conn.execute("""
        SELECT u.id, u.empresa_id, u.username, u.nombre, u.rol, u.force_password_change
        FROM usuarios u WHERE u.id=? AND u.activo=1
    """, (user_id,)).fetchone()
    conn.close()
    if row:
        return User(row['id'], row['empresa_id'], row['username'], row['nombre'], row['rol'],
                    row['force_password_change'] or 0)
    return None

# ════════════ HELPERS TENANT ════════════

def get_tenant_db_path(empresa_id):
    conn = get_master_db()
    row = conn.execute("SELECT db_path, slug FROM empresas WHERE id=? AND activa=1", (empresa_id,)).fetchone()
    conn.close()
    if row:
        return row['db_path'], row['slug']
    return None, None

def get_tenant_db(empresa_id):
    db_path, _ = get_tenant_db_path(empresa_id)
    if not db_path:
        return None
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_tenant_db(db_path):
    conn = sqlite3.connect(db_path)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS proveedores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            razon_social TEXT,
            moneda TEXT DEFAULT 'USD',
            plazo_dias INTEGER DEFAULT 30,
            nota_plazo TEXT,
            banco1_nombre TEXT, banco1_cuenta TEXT,
            banco2_nombre TEXT, banco2_cuenta TEXT,
            banco3_nombre TEXT, banco3_cuenta TEXT,
            email TEXT DEFAULT '',
            descuento_pct REAL DEFAULT 0,
            activo INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS facturas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proveedor_id INTEGER NOT NULL REFERENCES proveedores(id),
            numero TEXT NOT NULL,
            fecha DATE NOT NULL,
            fecha_vencimiento DATE NOT NULL,
            importe REAL NOT NULL,
            moneda TEXT DEFAULT 'USD',
            tipo_pago TEXT DEFAULT 'transferencia',
            cheque_id TEXT,
            estado TEXT DEFAULT 'pendiente',
            categoria TEXT DEFAULT 'empresa',
            nota TEXT,
            descuento REAL DEFAULT 0,
            destino TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS pagos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            factura_id INTEGER NOT NULL REFERENCES facturas(id),
            fecha_pago DATE NOT NULL,
            importe_pagado REAL NOT NULL,
            referencia TEXT,
            comprobante TEXT,
            nota TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS cobros_pagos (
            id INTEGER PRIMARY KEY,
            fecha DATE NOT NULL,
            tipo TEXT NOT NULL,
            cliente TEXT NOT NULL,
            moneda TEXT DEFAULT 'USD',
            monto REAL NOT NULL,
            medio TEXT DEFAULT 'Transferencia',
            banco TEXT DEFAULT '',
            nota TEXT DEFAULT '',
            comprobante TEXT DEFAULT '',
            estado TEXT DEFAULT 'Pendiente',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS saldos_favor (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proveedor_id INTEGER NOT NULL REFERENCES proveedores(id),
            fecha DATE NOT NULL,
            monto REAL NOT NULL,
            moneda TEXT DEFAULT 'USD',
            motivo TEXT,
            referencia TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)
    conn.commit()
    conn.close()

# Asegurar que exista la DB del admin
admin_db_path = os.path.join(TENANT_DIR, 'admin.db')
if not os.path.exists(admin_db_path):
    init_tenant_db(admin_db_path)

# ════════════ DECORATORS ════════════

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated

# ════════════ CONFIG POR TENANT ════════════

def load_config_tenant(empresa_id):
    defaults = {"carpeta_comprobantes": "", "nombre_empresa": "", "smtp_activo": False,
                "drive_folder_id": "", "google_credentials_file": ""}
    cfg_file = os.path.join(TENANT_DIR, f'config_{empresa_id}.json')
    if os.path.exists(cfg_file):
        try:
            with open(cfg_file) as f:
                data = json.load(f)
            defaults.update(data)
        except Exception:
            pass
    return defaults

def save_config_tenant(empresa_id, data):
    cfg_file = os.path.join(TENANT_DIR, f'config_{empresa_id}.json')
    existing = load_config_tenant(empresa_id)
    existing.update(data)
    with open(cfg_file, "w") as f:
        json.dump(existing, f, indent=2)

# ════════════ FUNCIONES EXISTENTES (adaptadas) ════════════

def buscar_comprobante(proveedor_nombre, fecha_pago_iso, empresa_id):
    cfg = load_config_tenant(empresa_id)
    carpeta_base = cfg.get("carpeta_comprobantes", "")
    if not carpeta_base or not os.path.isdir(carpeta_base):
        return []
    try:
        d = date.fromisoformat(fecha_pago_iso)
        subcarpeta = os.path.join(carpeta_base, str(d.year), f"{d.month:02d}")
        carpetas = [subcarpeta, carpeta_base] if os.path.isdir(subcarpeta) else [carpeta_base]
        prov_norm = proveedor_nombre.upper().replace(" ", "_").replace(",", "").replace(".", "")
        fecha_str = f"{d.day:02d}_{d.month:02d}_{d.year}"
        encontrados = []
        for carpeta in carpetas:
            if not os.path.isdir(carpeta):
                continue
            for fname in os.listdir(carpeta):
                fname_up = fname.upper().replace(" ", "_")
                prov_short = prov_norm[:6]
                if (prov_norm in fname_up or prov_short in fname_up) and fecha_str in fname_up:
                    ruta = os.path.join(carpeta, fname)
                    if ruta not in [r for r, _ in encontrados]:
                        encontrados.append((ruta, fname))
            if not encontrados:
                for fname in os.listdir(carpeta):
                    fname_up = fname.upper()
                    if fecha_str in fname_up and fname_up.startswith("PAGO_"):
                        ruta = os.path.join(carpeta, fname)
                        encontrados.append((ruta, fname))
        return encontrados
    except Exception:
        return []

def dias_para_vencer(fecha_str):
    if not fecha_str:
        return None
    try:
        fv = date.fromisoformat(str(fecha_str)) if isinstance(fecha_str, str) else fecha_str
        return (fv - date.today()).days
    except:
        return None

def get_resumen(empresa_id):
    db = get_tenant_db(empresa_id)
    if not db:
        return {}
    total_usd = db.execute("SELECT COALESCE(SUM(importe),0) FROM facturas WHERE estado='pendiente' AND moneda='USD'").fetchone()[0]
    total_uyu = db.execute("SELECT COALESCE(SUM(importe),0) FROM facturas WHERE estado='pendiente' AND moneda='UYU'").fetchone()[0]
    vencidas = db.execute("SELECT COUNT(*) FROM facturas WHERE estado='pendiente' AND fecha_vencimiento < date('now')").fetchone()[0]
    proximas = db.execute("SELECT COUNT(*) FROM facturas WHERE estado='pendiente' AND fecha_vencimiento BETWEEN date('now') AND date('now','+7 days')").fetchone()[0]
    total_cobros_mes = db.execute("SELECT COALESCE(SUM(monto),0) FROM cobros_pagos WHERE tipo='Cobro' AND strftime('%Y-%m',fecha)=strftime('%Y-%m','now')").fetchone()[0]
    total_pagos_mes = db.execute("SELECT COALESCE(SUM(monto),0) FROM cobros_pagos WHERE tipo='Pago' AND strftime('%Y-%m',fecha)=strftime('%Y-%m','now')").fetchone()[0]
    cp_pendientes = db.execute("SELECT COUNT(*) FROM cobros_pagos WHERE estado='Pendiente'").fetchone()[0]
    db.close()
    return {
        "total_usd": total_usd, "total_uyu": total_uyu,
        "vencidas": vencidas, "proximas": proximas,
        "cobros_mes": total_cobros_mes, "pagos_mes": total_pagos_mes,
        "cp_pendientes": cp_pendientes
    }

# ════════════ RUTAS AUTH ════════════

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if not username or not password:
            error = 'Completá todos los campos.'
        else:
            conn = get_master_db()
            row = conn.execute("""
                SELECT u.id, u.empresa_id, u.username, u.password_hash, u.nombre, u.rol,
                       u.activo, u.force_password_change,
                       e.nombre as empresa_nombre
                FROM usuarios u
                JOIN empresas e ON e.id = u.empresa_id
                WHERE u.username=?
            """, (username,)).fetchone()
            conn.close()

            if row and check_password_hash(row['password_hash'], password) and row['activo']:
                user = User(row['id'], row['empresa_id'], row['username'], row['nombre'], row['rol'],
                            row['force_password_change'] or 0)
                login_user(user)
                # Forzar cambio de contraseña si es primer login
                if user.force_password_change:
                    flash('Por seguridad, debés cambiar tu contraseña ahora.', 'warning')
                    return redirect(url_for('cambiar_password'))
                return redirect(url_for('index'))
            else:
                error = 'Usuario o contraseña incorrectos.'

    return render_template('login.html', error=error)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# ════════════ CAMBIO DE CONTRASEÑA (primer login / forzado) ════════════

@app.route('/cambiar-password', methods=['GET', 'POST'])
@login_required
def cambiar_password():
    if request.method == 'POST':
        new_pw = request.form.get('new_password', '').strip()
        confirm_pw = request.form.get('confirm_password', '').strip()

        if not new_pw or not confirm_pw:
            flash('Completá ambos campos.', 'danger')
            return redirect(url_for('cambiar_password'))
        if new_pw != confirm_pw:
            flash('Las contraseñas no coinciden.', 'danger')
            return redirect(url_for('cambiar_password'))
        if len(new_pw) < 8:
            flash('La contraseña debe tener al menos 8 caracteres.', 'danger')
            return redirect(url_for('cambiar_password'))

        conn = get_master_db()
        pw_hash = generate_password_hash(new_pw)
        conn.execute("UPDATE usuarios SET password_hash=?, force_password_change=0 WHERE id=?",
                     (pw_hash, current_user.id))
        conn.commit()
        conn.close()
        flash('Contraseña actualizada correctamente.', 'success')
        return redirect(url_for('index'))

    return render_template('cambiar_password.html')

# ════════════ RUTAS ADMIN (gestión de empresas y usuarios) ════════════

@app.route('/admin')
@login_required
@admin_required
def admin_panel():
    conn = get_master_db()
    empresas = conn.execute("""
        SELECT e.*, COUNT(u.id) as num_usuarios
        FROM empresas e
        LEFT JOIN usuarios u ON u.empresa_id = e.id AND u.activo = 1
        GROUP BY e.id ORDER BY e.nombre
    """).fetchall()
    usuarios = conn.execute("""
        SELECT u.*, e.nombre as empresa_nombre
        FROM usuarios u
        JOIN empresas e ON e.id = u.empresa_id
        ORDER BY e.nombre, u.username
    """).fetchall()
    conn.close()
    return render_template('admin.html', empresas=empresas, usuarios=usuarios)

@app.route('/admin/empresa/nueva', methods=['POST'])
@login_required
@admin_required
def admin_nueva_empresa():
    nombre = request.form.get('nombre', '').strip()
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    usuario_nombre = request.form.get('usuario_nombre', '').strip()

    if not nombre or not username or not password:
        flash('Todos los campos son obligatorios.', 'danger')
        return redirect(url_for('admin_panel'))

    slug = re.sub(r'[^a-z0-9]+', '-', nombre.lower()).strip('-')
    db_path = os.path.join(TENANT_DIR, f'{slug}.db')

    conn = get_master_db()
    # Verificar duplicados
    existing = conn.execute("SELECT id FROM empresas WHERE slug=?", (slug,)).fetchone()
    if existing:
        slug = f"{slug}-{secrets.token_hex(3)}"
        db_path = os.path.join(TENANT_DIR, f'{slug}.db')

    conn.execute(
        "INSERT INTO empresas (nombre, slug, db_path) VALUES (?,?,?)",
        (nombre, slug, db_path)
    )
    empresa_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    pw_hash = generate_password_hash(password)
    conn.execute(
        "INSERT INTO usuarios (empresa_id, username, password_hash, nombre, rol, force_password_change) VALUES (?,?,?,?,?,?)",
        (empresa_id, username, pw_hash, usuario_nombre or username, 'usuario', 1)
    )
    conn.commit()
    conn.close()

    # Crear DB del tenant
    init_tenant_db(db_path)

    flash(f'Empresa "{nombre}" creada. Usuario: {username} (debe cambiar contraseña en primer login)', 'success')
    return redirect(url_for('admin_panel'))

@app.route('/admin/empresa/<int:eid>/usuario/nuevo', methods=['POST'])
@login_required
@admin_required
def admin_nuevo_usuario(eid):
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    nombre = request.form.get('nombre', '').strip()

    if not username or not password:
        flash('Username y password obligatorios.', 'danger')
        return redirect(url_for('admin_panel'))

    conn = get_master_db()
    pw_hash = generate_password_hash(password)
    try:
        conn.execute(
            "INSERT INTO usuarios (empresa_id, username, password_hash, nombre, rol, force_password_change) VALUES (?,?,?,?,?,?)",
            (eid, username, pw_hash, nombre or username, 'usuario', 1)
        )
        conn.commit()
        flash(f'Usuario "{username}" creado (debe cambiar contraseña en primer login).', 'success')
    except Exception:
        flash(f'El usuario "{username}" ya existe en esa empresa.', 'danger')
    conn.close()
    return redirect(url_for('admin_panel'))

@app.route('/admin/empresa/<int:eid>/toggle', methods=['POST'])
@login_required
@admin_required
def admin_toggle_empresa(eid):
    conn = get_master_db()
    row = conn.execute("SELECT activa FROM empresas WHERE id=?", (eid,)).fetchone()
    if row:
        nueva = 0 if row['activa'] else 1
        conn.execute("UPDATE empresas SET activa=? WHERE id=?", (nueva, eid))
        conn.commit()
    conn.close()
    return redirect(url_for('admin_panel'))

@app.route('/admin/usuario/<int:uid>/reset-password', methods=['POST'])
@login_required
@admin_required
def admin_reset_password(uid):
    new_pw = request.form.get('new_password', '').strip()
    if not new_pw:
        flash('La contraseña no puede estar vacía.', 'danger')
        return redirect(url_for('admin_panel'))
    conn = get_master_db()
    pw_hash = generate_password_hash(new_pw)
    conn.execute("UPDATE usuarios SET password_hash=?, force_password_change=1 WHERE id=?", (pw_hash, uid))
    conn.commit()
    conn.close()
    flash('Contraseña actualizada.', 'success')
    return redirect(url_for('admin_panel'))

# ════════════ RUTAS PRINCIPALES (con login) ════════════

@app.route('/')
@login_required
def index():
    resumen = get_resumen(current_user.empresa_id)
    db = get_tenant_db(current_user.empresa_id)
    alertas = db.execute("""
        SELECT f.*, p.nombre as proveedor_nombre,
               julianday(f.fecha_vencimiento) - julianday('now') as dias
        FROM facturas f JOIN proveedores p ON p.id = f.proveedor_id
        WHERE f.estado = 'pendiente' AND f.fecha_vencimiento <= date('now','+7 days')
        ORDER BY f.fecha_vencimiento ASC
    """).fetchall()
    cp_recientes = db.execute("SELECT * FROM cobros_pagos ORDER BY fecha DESC, id DESC LIMIT 5").fetchall()
    db.close()
    return render_template('index.html', resumen=resumen, alertas=alertas,
                           cp_recientes=cp_recientes, today=date.today().isoformat())

# ════════════ PROVEEDORES ════════════

@app.route('/proveedores')
@login_required
def proveedores():
    db = get_tenant_db(current_user.empresa_id)
    rows = db.execute("""
        SELECT p.*,
            COUNT(f.id) as facturas_pendientes,
            COALESCE(SUM(CASE WHEN f.estado IN ('pendiente','aprobada') THEN f.importe ELSE 0 END), 0) as deuda_total,
            COALESCE((SELECT SUM(monto) FROM saldos_favor WHERE proveedor_id=p.id), 0) as saldo_favor
        FROM proveedores p
        LEFT JOIN facturas f ON f.proveedor_id=p.id AND f.estado IN ('pendiente','aprobada')
        WHERE p.activo=1
        GROUP BY p.id ORDER BY p.nombre
    """).fetchall()
    db.close()
    return render_template('proveedores.html', proveedores=rows)

@app.route('/proveedores/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_proveedor():
    if request.method == 'POST':
        f = request.form
        db = get_tenant_db(current_user.empresa_id)
        db.execute("""
            INSERT INTO proveedores (nombre, razon_social, moneda, plazo_dias, nota_plazo,
                banco1_nombre, banco1_cuenta, banco2_nombre, banco2_cuenta, banco3_nombre, banco3_cuenta,
                email, descuento_pct)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (f['nombre'], f.get('razon_social'), f.get('moneda','USD'),
              int(f.get('plazo_dias',30) or 30), f.get('nota_plazo'),
              f.get('banco1_nombre'), f.get('banco1_cuenta'),
              f.get('banco2_nombre'), f.get('banco2_cuenta'),
              f.get('banco3_nombre'), f.get('banco3_cuenta'),
              f.get('email',''), float(f.get('descuento_pct') or 0)))
        db.commit()
        db.close()
        return redirect(url_for('proveedores'))
    return render_template('proveedor_form.html', proveedor=None)

@app.route('/proveedores/<int:pid>/editar', methods=['GET', 'POST'])
@login_required
def editar_proveedor(pid):
    db = get_tenant_db(current_user.empresa_id)
    if request.method == 'POST':
        f = request.form
        db.execute("""
            UPDATE proveedores SET nombre=?, razon_social=?, moneda=?, plazo_dias=?, nota_plazo=?,
                banco1_nombre=?, banco1_cuenta=?, banco2_nombre=?, banco2_cuenta=?,
                banco3_nombre=?, banco3_cuenta=?, email=?, descuento_pct=?
            WHERE id=?
        """, (f['nombre'], f.get('razon_social'), f.get('moneda','USD'),
              int(f.get('plazo_dias',30) or 30), f.get('nota_plazo'),
              f.get('banco1_nombre'), f.get('banco1_cuenta'),
              f.get('banco2_nombre'), f.get('banco2_cuenta'),
              f.get('banco3_nombre'), f.get('banco3_cuenta'),
              f.get('email',''), float(f.get('descuento_pct') or 0), pid))
        db.commit()
        db.close()
        return redirect(url_for('proveedores'))
    p = db.execute("SELECT * FROM proveedores WHERE id=?", (pid,)).fetchone()
    db.close()
    return render_template('proveedor_form.html', proveedor=p)

# ════════════ FACTURAS ════════════

@app.route('/facturas')
@login_required
def facturas():
    estado = request.args.get('estado', 'pendiente')
    proveedor_id = request.args.get('proveedor_id', '')
    destino = request.args.get('destino', '')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    tipo_fecha = request.args.get('tipo_fecha', 'emision')

    db = get_tenant_db(current_user.empresa_id)
    query = """
        SELECT f.*, p.nombre as proveedor_nombre,
               julianday(f.fecha_vencimiento) - julianday('now') as dias
        FROM facturas f JOIN proveedores p ON p.id = f.proveedor_id
        WHERE 1=1
    """
    params = []
    if estado:
        query += " AND f.estado=?"; params.append(estado)
    if proveedor_id:
        query += " AND f.proveedor_id=?"; params.append(proveedor_id)
    if destino:
        query += " AND f.destino=?"; params.append(destino)
    campo_fecha = "f.fecha" if tipo_fecha == "emision" else "f.fecha_vencimiento"
    if fecha_desde:
        query += f" AND {campo_fecha} >= ?"; params.append(fecha_desde)
    if fecha_hasta:
        query += f" AND {campo_fecha} <= ?"; params.append(fecha_hasta)
    query += " ORDER BY f.fecha_vencimiento ASC"
    rows = db.execute(query, params).fetchall()
    provs = db.execute("SELECT id, nombre FROM proveedores WHERE activo=1 ORDER BY nombre").fetchall()
    db.close()
    return render_template('facturas.html', facturas=rows, proveedores=provs,
                           estado_filtro=estado, proveedor_filtro=proveedor_id,
                           destino_filtro=destino,
                           fecha_desde_filtro=fecha_desde,
                           fecha_hasta_filtro=fecha_hasta,
                           tipo_fecha_filtro=tipo_fecha)

@app.route('/facturas/nueva', methods=['GET', 'POST'])
@login_required
def nueva_factura():
    db = get_tenant_db(current_user.empresa_id)
    if request.method == 'POST':
        f = request.form
        fecha = f['fecha']
        proveedor_id = int(f['proveedor_id'])
        p = db.execute("SELECT * FROM proveedores WHERE id=?", (proveedor_id,)).fetchone()
        plazo = int(f.get('plazo_override') or p['plazo_dias'] or 30)
        fecha_obj = date.fromisoformat(fecha)
        fecha_vto = (fecha_obj + timedelta(days=plazo)).isoformat()
        descuento = float(f.get('descuento') or 0)
        db.execute("""
            INSERT INTO facturas (proveedor_id, numero, fecha, fecha_vencimiento, importe, moneda,
                tipo_pago, cheque_id, categoria, nota, descuento, destino)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (proveedor_id, f['numero'], fecha, fecha_vto,
              float(f['importe']), f.get('moneda','USD'),
              f.get('tipo_pago','transferencia'), f.get('cheque_id'),
              f.get('categoria','empresa'), f.get('nota'),
              descuento, f.get('destino','')))
        db.commit()
        db.close()
        return redirect(url_for('facturas'))
    provs = db.execute("SELECT id, nombre, moneda, plazo_dias, COALESCE(descuento_pct,0) as descuento_pct FROM proveedores WHERE activo=1 ORDER BY nombre").fetchall()
    db.close()
    return render_template('factura_form.html', factura=None, proveedores=provs, today=date.today().isoformat())

@app.route('/facturas/<int:fid>/pagar', methods=['GET', 'POST'])
@login_required
def pagar_factura(fid):
    db = get_tenant_db(current_user.empresa_id)
    factura = db.execute("""
        SELECT f.*, p.nombre as proveedor_nombre
        FROM facturas f JOIN proveedores p ON p.id=f.proveedor_id
        WHERE f.id=?
    """, (fid,)).fetchone()
    saldo_favor = db.execute(
        "SELECT COALESCE(SUM(monto),0) FROM saldos_favor WHERE proveedor_id=?",
        (factura['proveedor_id'],)
    ).fetchone()[0]

    if request.method == 'POST':
        f = request.form
        importe_pagado = float(f['importe_pagado'])
        usar_saldo = f.get('usar_saldo') == '1'
        if usar_saldo and saldo_favor > 0:
            saldo_usado = min(saldo_favor, float(factura['importe']))
            saldo_restante = saldo_favor - saldo_usado
            db.execute("DELETE FROM saldos_favor WHERE proveedor_id=?", (factura['proveedor_id'],))
            if saldo_restante > 0.001:
                db.execute("""
                    INSERT INTO saldos_favor (proveedor_id, fecha, monto, moneda, motivo)
                    VALUES (?, ?, ?, ?, 'Saldo restante tras aplicar a factura')
                """, (factura['proveedor_id'], f['fecha_pago'], round(saldo_restante, 2), factura['moneda']))
            nota_pago = (f.get('nota') or '') + f" [Saldo a favor aplicado: {saldo_usado:.2f} {factura['moneda']}]"
        else:
            nota_pago = f.get('nota')
        db.execute("""
            INSERT INTO pagos (factura_id, fecha_pago, importe_pagado, referencia, comprobante, nota)
            VALUES (?,?,?,?,?,?)
        """, (fid, f['fecha_pago'], importe_pagado,
              f.get('referencia'), f.get('comprobante'), nota_pago))
        db.execute("UPDATE facturas SET estado='pagada' WHERE id=?", (fid,))
        db.commit()
        pago_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        db.close()
        if f.get('generar_mail'):
            return redirect(url_for('generar_mail', pid=pago_id))
        return redirect(url_for('pagos'))

    db.close()
    return render_template('pagar_form.html', factura=factura,
                           saldo_favor=saldo_favor,
                           today=date.today().isoformat())

@app.route('/facturas/<int:fid>/estado', methods=['POST'])
@login_required
def cambiar_estado(fid):
    nuevo = request.form.get('estado')
    if nuevo in ('pendiente','aprobada','pagada','pagada_sin_comprobante','en_disputa'):
        db = get_tenant_db(current_user.empresa_id)
        db.execute("UPDATE facturas SET estado=? WHERE id=?", (nuevo, fid))
        db.commit()
        db.close()
    return redirect(request.referrer or url_for('facturas'))

@app.route('/facturas/<int:fid>/eliminar', methods=['POST'])
@login_required
def eliminar_factura(fid):
    forzar = request.form.get('forzar') == '1'
    db = get_tenant_db(current_user.empresa_id)
    factura = db.execute("SELECT * FROM facturas WHERE id=?", (fid,)).fetchone()
    if not factura:
        db.close()
        return redirect(url_for('facturas'))
    pagos_count = db.execute("SELECT COUNT(*) as n FROM pagos WHERE factura_id=?", (fid,)).fetchone()['n']
    if pagos_count > 0 and not forzar:
        db.close()
        return redirect(url_for('facturas') + f"?error_eliminar={fid}&pagos={pagos_count}")
    if pagos_count > 0 and forzar:
        db.execute("DELETE FROM pagos WHERE factura_id=?", (fid,))
    db.execute("DELETE FROM facturas WHERE id=?", (fid,))
    db.commit()
    db.close()
    return redirect(request.referrer or url_for('facturas'))

@app.route('/facturas/duplicados')
@login_required
def facturas_duplicados():
    def norm(numero):
        m = re.findall(r'\d{5,}', str(numero))
        return m[-1] if m else numero.strip().lower()

    db = get_tenant_db(current_user.empresa_id)
    todas = db.execute("""
        SELECT f.*, p.nombre as proveedor_nombre
        FROM facturas f JOIN proveedores p ON p.id = f.proveedor_id
        ORDER BY f.proveedor_id, f.importe, f.fecha
    """).fetchall()
    db.close()

    grupos_numero = {}
    for f in todas:
        key = (f['proveedor_id'], norm(f['numero']))
        grupos_numero.setdefault(key, []).append(dict(f))

    grupos_extra = []
    lista = [dict(f) for f in todas]
    usados = set()
    for i, a in enumerate(lista):
        if a['id'] in usados: continue
        grupo = [a]
        try: fa = date.fromisoformat(a['fecha'])
        except: continue
        for j, b in enumerate(lista):
            if i == j or b['id'] in usados: continue
            if a['proveedor_id'] != b['proveedor_id']: continue
            if abs(a['importe'] - b['importe']) > 0.01: continue
            if norm(a['numero']) == norm(b['numero']): continue
            try: fb = date.fromisoformat(b['fecha'])
            except: continue
            if abs((fa - fb).days) <= 7:
                grupo.append(b)
        if len(grupo) > 1:
            for g in grupo: usados.add(g['id'])
            grupos_extra.append(grupo)

    duplicados = []
    for key, grupo in grupos_numero.items():
        if len(grupo) > 1:
            duplicados.append({"tipo": "numero", "grupo": grupo})
    for grupo in grupos_extra:
        duplicados.append({"tipo": "importe", "grupo": grupo})

    return render_template('duplicados.html', duplicados=duplicados, total=len(duplicados))

# ════════════ PAGOS ════════════

@app.route('/pagos')
@login_required
def pagos():
    proveedor_id = request.args.get('proveedor_id', '')
    db = get_tenant_db(current_user.empresa_id)
    query = """
        SELECT pg.*, f.numero as factura_numero, f.importe as factura_importe,
               f.moneda, f.fecha as factura_fecha, p.nombre as proveedor_nombre, p.razon_social
        FROM pagos pg JOIN facturas f ON f.id = pg.factura_id
        JOIN proveedores p ON p.id = f.proveedor_id WHERE 1=1
    """
    params = []
    if proveedor_id:
        query += " AND p.id = ?"; params.append(proveedor_id)
    query += " ORDER BY pg.fecha_pago DESC, pg.id DESC"
    rows = db.execute(query, params).fetchall()
    provs = db.execute("SELECT id, nombre FROM proveedores WHERE activo=1 ORDER BY nombre").fetchall()
    db.close()
    from collections import OrderedDict
    grupos = OrderedDict()
    for r in rows:
        key = (r['proveedor_nombre'], r['fecha_pago'], r['comprobante'] or r['referencia'] or str(r['id']))
        if key not in grupos:
            grupos[key] = {"info": r, "facturas": [], "total": 0}
        grupos[key]["facturas"].append(r); grupos[key]["total"] += r['importe_pagado']
    return render_template('pagos.html', grupos=grupos, proveedores=provs,
                           proveedor_filtro=proveedor_id, today=date.today().isoformat())

@app.route('/pagos/<int:pid>/revertir', methods=['POST'])
@login_required
def revertir_pago(pid):
    db = get_tenant_db(current_user.empresa_id)
    pago = db.execute("SELECT * FROM pagos WHERE id=?", (pid,)).fetchone()
    if not pago:
        db.close()
        return "Pago no encontrado", 404
    db.execute("UPDATE facturas SET estado='pendiente' WHERE id=?", (pago['factura_id'],))
    db.execute("DELETE FROM pagos WHERE id=?", (pid,))
    db.commit()
    db.close()
    return redirect(request.referrer or url_for('pagos'))

@app.route('/pagos/<int:pid>/editar', methods=['GET', 'POST'])
@login_required
def editar_pago(pid):
    db = get_tenant_db(current_user.empresa_id)
    pago = db.execute("""
        SELECT pg.*, f.numero as factura_numero, f.importe as factura_importe,
               f.moneda, p.nombre as proveedor_nombre
        FROM pagos pg JOIN facturas f ON f.id = pg.factura_id
        JOIN proveedores p ON p.id = f.proveedor_id WHERE pg.id=?
    """, (pid,)).fetchone()
    if not pago:
        db.close()
        return "Pago no encontrado", 404
    if request.method == 'POST':
        f = request.form
        db.execute("UPDATE pagos SET fecha_pago=?, importe_pagado=?, referencia=?, comprobante=?, nota=? WHERE id=?",
            (f['fecha_pago'], float(f['importe_pagado']), f.get('referencia',''), f.get('comprobante',''), f.get('nota',''), pid))
        db.commit()
        db.close()
        return redirect(url_for('pagos'))
    db.close()
    return render_template('editar_pago.html', pago=pago, today=date.today().isoformat())

@app.route('/pagos/<int:pid>/mail')
@login_required
def generar_mail(pid):
    db = get_tenant_db(current_user.empresa_id)
    pago = db.execute("SELECT * FROM pagos WHERE id=?", (pid,)).fetchone()
    if not pago:
        db.close()
        return "Pago no encontrado", 404
    factura = db.execute("""
        SELECT f.*, p.nombre as proveedor_nombre, p.razon_social,
               COALESCE(p.email, '') as proveedor_email
        FROM facturas f JOIN proveedores p ON p.id=f.proveedor_id WHERE f.id=?
    """, (pago['factura_id'],)).fetchone()
    facturas_batch = db.execute("""
        SELECT DISTINCT f.numero, f.importe, f.moneda
        FROM pagos pg JOIN facturas f ON f.id=pg.factura_id
        WHERE f.proveedor_id=? AND pg.fecha_pago=? ORDER BY f.fecha
    """, (factura['proveedor_id'], pago['fecha_pago'])).fetchall()
    db.close()
    archivos_encontrados = buscar_comprobante(factura['proveedor_nombre'], pago['fecha_pago'], current_user.empresa_id)
    cfg = load_config_tenant(current_user.empresa_id)
    return render_template('mail_pago.html', pago=pago, factura=factura,
                           facturas_batch=facturas_batch, archivos_encontrados=archivos_encontrados,
                           carpeta_ok=bool(cfg.get('carpeta_comprobantes')), nombre_empresa=cfg.get('nombre_empresa',''))

@app.route('/pagos/<int:pid>/descargar_eml', methods=['POST'])
@login_required
def descargar_eml(pid):
    import email as emaillib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    import mimetypes
    db = get_tenant_db(current_user.empresa_id)
    pago = db.execute("SELECT * FROM pagos WHERE id=?", (pid,)).fetchone()
    if not pago:
        db.close()
        return "Pago no encontrado", 404
    factura = db.execute("""
        SELECT f.*, p.nombre as proveedor_nombre, p.razon_social, COALESCE(p.email,'') as proveedor_email
        FROM facturas f JOIN proveedores p ON p.id=f.proveedor_id WHERE f.id=?
    """, (pago['factura_id'],)).fetchone()
    facturas_batch = db.execute("""
        SELECT DISTINCT f.numero, f.importe, f.moneda FROM pagos pg JOIN facturas f ON f.id=pg.factura_id
        WHERE f.proveedor_id=? AND pg.fecha_pago=? ORDER BY f.fecha
    """, (factura['proveedor_id'], pago['fecha_pago'])).fetchall()
    db.close()
    cfg = load_config_tenant(current_user.empresa_id)
    nombre_empresa = cfg.get('nombre_empresa','')
    archivo_path = request.form.get('archivo_path','').strip()
    tono = request.form.get('tono','formal')
    lista = "\n".join(f"  • {f['numero']} — {f['moneda']} {f['importe']:.2f}" for f in facturas_batch)
    total = sum(f['importe'] for f in facturas_batch)
    moneda = facturas_batch[0]['moneda'] if facturas_batch else 'USD'
    ref_line = f"Referencia: {pago['referencia']}\n" if pago['referencia'] else ""
    if tono == 'breve':
        cuerpo = f"Hola,\n\nLes informamos que realizamos el pago de {moneda} {total:.2f} con fecha {pago['fecha_pago']}.\n\nFacturas abonadas:\n{lista}\n\n{ref_line}Adjunto comprobante.\n\nSaludos,\n{nombre_empresa}"
    else:
        cuerpo = f"Estimados,\n\nPor medio del presente les comunicamos que hemos efectuado el pago correspondiente a las siguientes facturas:\n\n{lista}\n\nTotal abonado: {moneda} {total:.2f}\nFecha de pago: {pago['fecha_pago']}\n{ref_line}\nAdjuntamos el comprobante de transferencia para su registro.\n\nSaludos cordiales,\n{nombre_empresa}"
    asunto = f"Comprobante de pago — {facturas_batch[0]['numero']}" if len(facturas_batch)==1 else f"Comprobante de pago — {factura['proveedor_nombre']} ({len(facturas_batch)} facturas)"
    msg = MIMEMultipart()
    msg['From'] = ''; msg['To'] = factura['proveedor_email'] or ''; msg['Subject'] = asunto
    msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))
    if archivo_path and os.path.isfile(archivo_path):
        mime_type, _ = mimetypes.guess_type(archivo_path)
        main_type, sub_type = (mime_type or 'application/octet-stream').split('/', 1)
        with open(archivo_path, 'rb') as f:
            part = MIMEBase(main_type, sub_type); part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(archivo_path))
        msg.attach(part)
    out = io.BytesIO(msg.as_bytes()); out.seek(0)
    return send_file(out, download_name=f"pago_{factura['proveedor_nombre'].replace(' ','_')}_{pago['fecha_pago']}.eml", mimetype='message/rfc822')

# ════════════ INFORME ════════════

@app.route('/informe')
@login_required
def informe():
    destino = request.args.get('destino', '')
    db = get_tenant_db(current_user.empresa_id)
    query = """
        SELECT f.*, p.nombre as proveedor_nombre, p.razon_social,
               p.banco1_nombre, p.banco1_cuenta, p.banco2_nombre, p.banco2_cuenta,
               p.banco3_nombre, p.banco3_cuenta, p.moneda as moneda_prov,
               julianday(f.fecha_vencimiento) - julianday('now') as dias
        FROM facturas f JOIN proveedores p ON p.id = f.proveedor_id
        WHERE f.estado IN ('pendiente','aprobada')
    """
    params = []
    if destino:
        query += " AND f.destino=?"; params.append(destino)
    query += " ORDER BY p.nombre, f.fecha_vencimiento"
    rows = db.execute(query, params).fetchall()
    db.close()
    proveedores_data = {}
    for r in rows:
        pid = r['proveedor_id']
        if pid not in proveedores_data:
            proveedores_data[pid] = {"info": r, "facturas": [], "total": 0, "total_descuento": 0, "total_neto": 0}
        desc = float(r['descuento'] or 0); importe = float(r['importe'])
        monto_desc = importe * desc / 100 if desc > 0 else 0
        proveedores_data[pid]["facturas"].append(r)
        proveedores_data[pid]["total"] += importe
        proveedores_data[pid]["total_descuento"] += monto_desc
        proveedores_data[pid]["total_neto"] += importe - monto_desc
    return render_template('informe.html', proveedores=proveedores_data,
                           today=date.today().isoformat(), destino_filtro=destino)

@app.route('/informe/exportar')
@login_required
def exportar_informe():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    destino = request.args.get('destino', '')
    db = get_tenant_db(current_user.empresa_id)
    query = """
        SELECT f.*, p.nombre as proveedor_nombre, p.razon_social,
               p.banco1_nombre, p.banco1_cuenta, p.banco2_nombre, p.banco2_cuenta,
               p.banco3_nombre, p.banco3_cuenta
        FROM facturas f JOIN proveedores p ON p.id = f.proveedor_id
        WHERE f.estado IN ('pendiente','aprobada')
    """
    params = []
    if destino:
        query += " AND f.destino=?"; params.append(destino)
    query += " ORDER BY p.nombre, f.fecha"
    rows = db.execute(query, params).fetchall()
    db.close()
    wb = Workbook(); ws = wb.active; ws.title = "Pago a proveedores"
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    sub_font = Font(name="Arial", bold=True, size=9)
    data_font = Font(name="Arial", size=9)
    bold_font = Font(name="Arial", bold=True, size=9)
    blue_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    yellow_fill = PatternFill("solid", start_color="FFD700", end_color="FFD700")
    total_fill = PatternFill("solid", start_color="D9D9D9", end_color="D9D9D9")
    neto_fill = PatternFill("solid", start_color="D4EDDA", end_color="D4EDDA")
    desc_fill = PatternFill("solid", start_color="FFF3CD", end_color="FFF3CD")
    center = Alignment(horizontal="center", vertical="center")
    from collections import OrderedDict
    grupos = OrderedDict()
    for r in rows:
        k = r['proveedor_nombre']
        if k not in grupos:
            grupos[k] = {"rows": [], "info": r}
        grupos[k]["rows"].append(r)
    row_idx = 1; today_str = date.today().strftime("%d/%m/%Y")
    for nombre, grupo in grupos.items():
        info = grupo['info']; facturas_grp = grupo['rows']
        moneda = facturas_grp[0]['moneda'] if facturas_grp else 'USD'
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        c = ws.cell(row_idx, 1, nombre.upper()); c.font = hdr_font; c.fill = blue_fill; c.alignment = center
        ws.cell(row_idx, 5, today_str).font = hdr_font; ws.cell(row_idx, 5).fill = blue_fill
        ws.cell(row_idx, 9, "CUENTAS BANCARIAS").font = bold_font
        row_idx += 1
        bank_row = row_idx
        if info['razon_social']:
            ws.cell(bank_row, 9, "Titular:"); ws.cell(bank_row, 10, info['razon_social']).font = data_font; bank_row += 1
        for i in range(1, 4):
            bn = info[f'banco{i}_nombre']; bc = info[f'banco{i}_cuenta']
            if bn:
                ws.cell(bank_row, 9, bn).font = bold_font; ws.cell(bank_row, 10, bc or '').font = data_font; bank_row += 1
        headers = ["FECHA","N° FACTURA","IMPORTE","DESTINO","VTO CHEQUE","IMPORTE"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row_idx, col, h); c.font = sub_font; c.fill = yellow_fill; c.alignment = center
        row_idx += 1; data_start = row_idx; total_bruto = 0; total_desc_mto = 0
        for f in facturas_grp:
            importe = float(f['importe']); descuento = float(f['descuento'] or 0)
            desc_mto = importe * descuento / 100 if descuento > 0 else 0
            total_bruto += importe; total_desc_mto += desc_mto
            try:
                d = date.fromisoformat(str(f['fecha'])); ws.cell(row_idx, 1, d.strftime("%d/%m/%Y")).font = data_font
            except:
                ws.cell(row_idx, 1, f['fecha']).font = data_font
            ws.cell(row_idx, 2, f['numero']).font = data_font
            imp_c = ws.cell(row_idx, 3, importe); imp_c.font = data_font; imp_c.number_format = '#,##0.00'
            destino_val = f['destino'] or f['cheque_id'] or ''
            if destino_val: ws.cell(row_idx, 4, destino_val).font = data_font
            if f['nota']: ws.cell(row_idx, 7, f['nota']).font = Font(name="Arial", size=8, italic=True)
            row_idx += 1
        data_end = row_idx - 1
        ws.cell(row_idx, 1, "SUBTOTAL").font = bold_font; ws.cell(row_idx, 1).fill = total_fill
        sub_c = ws.cell(row_idx, 3, f"=SUM(C{data_start}:C{data_end})")
        sub_c.font = bold_font; sub_c.fill = total_fill; sub_c.number_format = '#,##0.00'
        row_idx += 1
        if total_desc_mto > 0:
            pct = round(total_desc_mto / total_bruto * 100, 1) if total_bruto else 0
            ws.cell(row_idx, 1, f"DESCUENTO ({pct}%)").font = Font(name="Arial", bold=True, size=9, color="9A3412")
            ws.cell(row_idx, 1).fill = desc_fill
            ws.cell(row_idx, 3, -round(total_desc_mto, 2)).number_format = '#,##0.00'; ws.cell(row_idx, 3).fill = desc_fill
            row_idx += 1
            ws.cell(row_idx, 1, "TOTAL A PAGAR").font = Font(name="Arial", bold=True, size=9, color="166534")
            ws.cell(row_idx, 1).fill = neto_fill
            ws.cell(row_idx, 3, round(total_bruto - total_desc_mto, 2)).number_format = '#,##0.00'; ws.cell(row_idx, 3).fill = neto_fill
            row_idx += 1
        row_idx += 1
    for col, w in [("A",14),("B",28),("C",13),("D",10),("E",12),("F",13),("G",18),("I",14),("J",30)]:
        ws.column_dimensions[col].width = w
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, download_name=f"Pago_Proveedores_{date.today().strftime('%Y%m%d')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ════════════ COBROS Y PAGOS ════════════

@app.route('/cobros-pagos')
@login_required
def cobros_pagos_registrar():
    return render_template('cobros_registrar.html', today=date.today().isoformat())

@app.route('/cobros-pagos/historial')
@login_required
def cobros_pagos_historial():
    db = get_tenant_db(current_user.empresa_id)
    rows = db.execute("SELECT * FROM cobros_pagos ORDER BY fecha DESC, id DESC").fetchall()
    db.close()
    return render_template('cobros_historial.html', registros=rows)

@app.route('/cobros-pagos/informe')
@login_required
def cobros_pagos_informe():
    db = get_tenant_db(current_user.empresa_id)
    rows = db.execute("SELECT * FROM cobros_pagos ORDER BY fecha DESC, id DESC").fetchall()
    db.close()
    total_cobros_usd = sum(r['monto'] for r in rows if r['tipo']=='Cobro' and r['moneda']=='USD')
    total_cobros_uyu = sum(r['monto'] for r in rows if r['tipo']=='Cobro' and r['moneda']=='UYU')
    total_pagos_usd = sum(r['monto'] for r in rows if r['tipo']=='Pago' and r['moneda']=='USD')
    total_pagos_uyu = sum(r['monto'] for r in rows if r['tipo']=='Pago' and r['moneda']=='UYU')
    return render_template('cobros_informe.html', registros=rows,
        total_cobros_usd=total_cobros_usd, total_cobros_uyu=total_cobros_uyu,
        total_pagos_usd=total_pagos_usd, total_pagos_uyu=total_pagos_uyu,
        today=date.today().isoformat())

# ════════════ API COBROS Y PAGOS ════════════

@app.route('/api/cobros_pagos', methods=['GET'])
@login_required
def api_cobros_pagos_list():
    tipo=request.args.get('tipo',''); estado=request.args.get('estado','')
    moneda=request.args.get('moneda',''); cliente=request.args.get('cliente','')
    desde=request.args.get('desde',''); hasta=request.args.get('hasta','')
    query="SELECT * FROM cobros_pagos WHERE 1=1"; params=[]
    if tipo: query+=" AND tipo=?"; params.append(tipo)
    if estado: query+=" AND estado=?"; params.append(estado)
    if moneda: query+=" AND moneda=?"; params.append(moneda)
    if cliente: query+=" AND cliente LIKE ?"; params.append(f'%{cliente}%')
    if desde: query+=" AND fecha>=?"; params.append(desde)
    if hasta: query+=" AND fecha<=?"; params.append(hasta)
    query+=" ORDER BY fecha DESC, id DESC"
    db = get_tenant_db(current_user.empresa_id)
    rows = db.execute(query, params).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/cobros_pagos', methods=['POST'])
@login_required
def api_cobros_pagos_create():
    data = request.get_json()
    if not data: return jsonify({"error":"Sin datos"}), 400
    cliente = (data.get('cliente') or '').strip()
    monto = float(data.get('monto') or 0)
    if not cliente or monto <= 0: return jsonify({"error":"Cliente y monto requeridos"}), 400
    record_id = data.get('id') or int(date.today().strftime('%Y%m%d') + str(abs(hash(cliente + str(monto))))[0:6])
    fecha = data.get('fecha') or date.today().isoformat()
    db = get_tenant_db(current_user.empresa_id)
    db.execute("INSERT OR IGNORE INTO cobros_pagos (id,fecha,tipo,cliente,moneda,monto,medio,banco,nota,comprobante,estado) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (record_id, fecha, data.get('tipo','Cobro'), cliente, data.get('moneda','USD'), monto,
         data.get('medio','Transferencia'), data.get('banco',''), data.get('nota',''),
         data.get('comprobante',''), data.get('estado','Pendiente')))
    db.commit()
    new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0] or record_id
    db.close()
    return jsonify({"id": new_id, "ok": True}), 201

@app.route('/api/cobros_pagos/<int:rid>', methods=['PUT'])
@login_required
def api_cobros_pagos_update(rid):
    data = request.get_json()
    if not data: return jsonify({"error":"Sin datos"}), 400
    db = get_tenant_db(current_user.empresa_id)
    if not db.execute("SELECT id FROM cobros_pagos WHERE id=?",(rid,)).fetchone():
        db.close(); return jsonify({"error":"No encontrado"}), 404
    db.execute("UPDATE cobros_pagos SET fecha=?,tipo=?,cliente=?,moneda=?,monto=?,medio=?,banco=?,nota=?,comprobante=?,estado=? WHERE id=?",
        (data.get('fecha'), data.get('tipo'), data.get('cliente'), data.get('moneda','USD'),
         float(data.get('monto') or 0), data.get('medio','Transferencia'), data.get('banco',''),
         data.get('nota',''), data.get('comprobante',''), data.get('estado','Pendiente'), rid))
    db.commit(); db.close()
    return jsonify({"ok":True})

@app.route('/api/cobros_pagos/<int:rid>/estado', methods=['POST'])
@login_required
def api_cobros_pagos_estado(rid):
    data = request.get_json()
    estado = data.get('estado') if data else None
    if estado not in ('Pendiente','Procesado'): return jsonify({"error":"Estado inválido"}), 400
    db = get_tenant_db(current_user.empresa_id)
    db.execute("UPDATE cobros_pagos SET estado=? WHERE id=?",(estado,rid))
    db.commit(); db.close()
    return jsonify({"ok":True})

@app.route('/api/cobros_pagos/<int:rid>', methods=['DELETE'])
@login_required
def api_cobros_pagos_delete(rid):
    db = get_tenant_db(current_user.empresa_id)
    db.execute("DELETE FROM cobros_pagos WHERE id=?",(rid,))
    db.commit(); db.close()
    return jsonify({"ok":True})

@app.route('/api/cobros_pagos/importar', methods=['POST'])
@login_required
def api_cobros_pagos_importar():
    data = request.get_json()
    registros = data.get('registros',[]) if data else []
    inserted=0; skipped=0
    db = get_tenant_db(current_user.empresa_id)
    for r in registros:
        try:
            db.execute("INSERT OR IGNORE INTO cobros_pagos (id,fecha,tipo,cliente,moneda,monto,medio,banco,nota,comprobante,estado) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (r['id'], r['fecha'], r['tipo'], r['cliente'], r.get('moneda','USD'), float(r.get('monto',0)),
                 r.get('medio','Transferencia'), r.get('banco',''), r.get('nota',''), r.get('comprobante',''), r.get('estado','Pendiente')))
            if db.execute("SELECT changes()").fetchone()[0]: inserted += 1
            else: skipped += 1
        except Exception: skipped += 1
    db.commit(); db.close()
    return jsonify({"insertados":inserted,"omitidos":skipped})

@app.route('/api/cobros_pagos/backup', methods=['GET'])
@login_required
def api_cobros_pagos_backup():
    db = get_tenant_db(current_user.empresa_id)
    rows = db.execute("SELECT * FROM cobros_pagos ORDER BY fecha DESC, id DESC").fetchall()
    db.close()
    registros = [dict(r) for r in rows]
    for r in registros: r.pop('created_at', None)
    return jsonify({"version":1,"exportado":date.today().isoformat()+"T00:00:00.000Z","registros":registros})

# ════════════ API PROVEEDORES ════════════

@app.route('/api/proveedores')
@login_required
def api_proveedores():
    db = get_tenant_db(current_user.empresa_id)
    rows = db.execute("SELECT id, nombre FROM proveedores WHERE activo=1 ORDER BY nombre").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/facturas_pendientes/<int:pid>')
@login_required
def api_facturas_pendientes(pid):
    db = get_tenant_db(current_user.empresa_id)
    rows = db.execute("""
        SELECT id, numero, importe, moneda, fecha_vencimiento,
               COALESCE(descuento,0) as descuento_fact
        FROM facturas WHERE proveedor_id=? AND estado IN ('pendiente','aprobada')
        ORDER BY fecha_vencimiento
    """, (pid,)).fetchall()
    prov = db.execute("SELECT COALESCE(descuento_pct,0) as descuento_pct FROM proveedores WHERE id=?", (pid,)).fetchone()
    db.close()
    return jsonify({"facturas": [dict(r) for r in rows], "descuento_pct": float(prov['descuento_pct']) if prov else 0})

@app.route('/api/marcar_pagadas', methods=['POST'])
@login_required
def api_marcar_pagadas():
    data = request.get_json()
    pid = data.get('proveedor_id')
    estado = data.get('estado', 'pagada')
    comprobante = data.get('comprobante', '')
    nota = data.get('nota', '')
    fecha_pago = data.get('fecha_pago') or date.today().isoformat()
    factura_ids = data.get('factura_ids', [])
    descuento_pct = float(data.get('descuento_pct', 0) or 0)
    if estado not in ('pagada', 'pagada_sin_comprobante'):
        estado = 'pagada'
    actualizadas = 0; total_pagado = 0
    db = get_tenant_db(current_user.empresa_id)
    if factura_ids:
        placeholders = ",".join("?" * len(factura_ids))
        facturas = db.execute(
            f"SELECT id, importe FROM facturas WHERE id IN ({placeholders}) AND proveedor_id=? AND estado IN ('pendiente','aprobada')",
            (*factura_ids, pid)
        ).fetchall()
    else:
        facturas = db.execute(
            "SELECT id, importe FROM facturas WHERE proveedor_id=? AND estado IN ('pendiente','aprobada')", (pid,)
        ).fetchall()
    for f in facturas:
        importe_original = float(f['importe'])
        importe_pagado = round(importe_original * (1 - descuento_pct / 100), 2) if descuento_pct > 0 else importe_original
        db.execute("UPDATE facturas SET estado=? WHERE id=?", (estado, f['id']))
        db.execute("INSERT INTO pagos (factura_id, fecha_pago, importe_pagado, comprobante, nota) VALUES (?,?,?,?,?)",
            (f['id'], fecha_pago, importe_pagado, comprobante, nota))
        actualizadas += 1; total_pagado += importe_pagado
    db.commit(); db.close()
    return jsonify({"actualizadas": actualizadas, "total_pagado": round(total_pagado, 2)})

@app.route('/api/facturas_pagadas/<int:pid>')
@login_required
def api_facturas_pagadas(pid):
    db = get_tenant_db(current_user.empresa_id)
    rows = db.execute("""
        SELECT pg.id as pago_id, pg.fecha_pago, pg.importe_pagado,
               f.id as factura_id, f.numero, f.importe, f.moneda
        FROM pagos pg JOIN facturas f ON f.id = pg.factura_id
        WHERE f.proveedor_id=? AND f.estado IN ('pagada','pagada_sin_comprobante')
        ORDER BY pg.fecha_pago DESC, pg.id DESC
    """, (pid,)).fetchall()
    prov = db.execute("SELECT COALESCE(descuento_pct,0) as descuento_pct FROM proveedores WHERE id=?", (pid,)).fetchone()
    db.close()
    return jsonify({"facturas": [dict(r) for r in rows], "descuento_pct": float(prov['descuento_pct']) if prov else 0})

@app.route('/api/revertir_pagos', methods=['POST'])
@login_required
def api_revertir_pagos():
    data = request.get_json()
    pago_ids = data.get('pago_ids', [])
    revertidos = 0
    db = get_tenant_db(current_user.empresa_id)
    for pago_id in pago_ids:
        pago = db.execute("SELECT * FROM pagos WHERE id=?", (pago_id,)).fetchone()
        if pago:
            db.execute("UPDATE facturas SET estado='pendiente' WHERE id=?", (pago['factura_id'],))
            db.execute("DELETE FROM pagos WHERE id=?", (pago_id,))
            revertidos += 1
    db.commit(); db.close()
    return jsonify({"revertidos": revertidos})

# ════════════ IMPORTAR / ESTADO CUENTA ════════════

@app.route('/importar', methods=['GET', 'POST'])
@login_required
def importar():
    resultado = None
    if request.method == 'POST':
        f = request.files.get('archivo')
        if not f: return render_template('importar.html', resultado={"error": "No se subió ningún archivo."})
        resultado = importar_excel(f, current_user.empresa_id)
    return render_template('importar.html', resultado=resultado)

def importar_excel(fileobj, empresa_id):
    from openpyxl import load_workbook
    SKIP_NUMEROS = {"FECHA","N° FACTURA","IMPORTE","TOTAL","TOTALES","PESOS","DOLARES","CUENTAS BANCARIAS","DESCUENTO 20%","DESCUENTO","SALDO DIFERENCIA"}
    try: wb = load_workbook(fileobj, data_only=True)
    except Exception as e: return {"error": f"No se pudo leer el Excel: {e}"}
    ws = wb.active
    db = get_tenant_db(empresa_id)
    provs = db.execute("SELECT id, nombre, plazo_dias FROM proveedores").fetchall()
    prov_map = {p['nombre'].upper().strip(): (p['id'], p['plazo_dias']) for p in provs}
    imported = 0; skipped = 0; not_found = []; errors = []
    current_prov_id = None; current_moneda = "USD"; current_plazo = 30; in_facturas = False
    for row in ws.iter_rows(values_only=True):
        a=row[0]; b=row[1] if len(row)>1 else None; c=row[2] if len(row)>2 else None
        d=row[3] if len(row)>3 else None; e=row[4] if len(row)>4 else None
        f_col=row[5] if len(row)>5 else None; g_col=row[6] if len(row)>6 else None
        if (isinstance(a,str) and a.strip() and b is None and c is None and isinstance(e,str) and e.strip() and isinstance(g_col,str) and g_col.strip() in ("USD","UYU")):
            current_prov_nombre=a.strip().upper(); current_moneda=g_col.strip(); in_facturas=False
            match=prov_map.get(current_prov_nombre)
            if match: current_prov_id,current_plazo=match
            else:
                current_prov_id=None
                for k,v in prov_map.items():
                    if current_prov_nombre in k or k in current_prov_nombre: current_prov_id,current_plazo=v; break
                if not current_prov_id and current_prov_nombre not in not_found: not_found.append(current_prov_nombre)
            continue
        if (isinstance(b,str) and b.strip() and c is None and d is None and hasattr(e,'date') and isinstance(f_col,str) and f_col.strip() in ("USD","UYU")):
            current_prov_nombre=b.strip().upper(); current_moneda=f_col.strip(); in_facturas=False
            match=prov_map.get(current_prov_nombre)
            if match: current_prov_id,current_plazo=match
            else:
                current_prov_id=None
                for k,v in prov_map.items():
                    if current_prov_nombre in k or k in current_prov_nombre: current_prov_id,current_plazo=v; break
                if not current_prov_id and current_prov_nombre not in not_found: not_found.append(current_prov_nombre)
            continue
        fecha_col_val=a or b
        if isinstance(fecha_col_val,str) and fecha_col_val.strip().upper()=="FECHA": in_facturas=True; continue
        if not in_facturas or not current_prov_id: continue
        fecha_val=a if (a and not isinstance(a,str)) or (isinstance(a,str) and a.strip() and a.strip() not in ("","-")) else b
        numero_val=b if isinstance(b,str) and b.strip() else (c if isinstance(c,str) and c.strip() else None)
        importe_val=c if isinstance(c,(int,float)) else (d if isinstance(d,(int,float)) else None)
        destino_val=d if isinstance(d,str) else (e if isinstance(e,str) else None)
        if numero_val is None or importe_val is None: continue
        numero=numero_val.strip()
        if numero.upper() in SKIP_NUMEROS or numero.upper().startswith("TOTAL") or numero.upper().startswith("DESCUENTO"): continue
        try: importe=float(importe_val)
        except: continue
        if hasattr(fecha_val,'date'): fecha_iso=fecha_val.date().isoformat()
        elif isinstance(fecha_val,str) and fecha_val.strip() not in ("-",""):
            try:
                from datetime import datetime
                fecha_iso=datetime.strptime(fecha_val.strip(),"%d/%m/%Y").date().isoformat()
            except: fecha_iso=date.today().isoformat()
        else: fecha_iso=date.today().isoformat()
        cheque_str=str(destino_val).strip().upper() if destino_val else ""
        if cheque_str in ("EVO","EVOLUCION","EVOLUCIÓN"): tipo_pago="cheque";cheque_id=cheque_str;categoria="empresa";destino_db="EVO"
        elif cheque_str in ("FEDE","FEDERICO"): tipo_pago="cheque";cheque_id=cheque_str;categoria="personal";destino_db="FEDE"
        elif cheque_str and cheque_str not in ("NONE",""): tipo_pago="cheque";cheque_id=cheque_str;categoria="empresa";destino_db=cheque_str
        else: tipo_pago="transferencia";cheque_id=None;categoria="empresa";destino_db=""
        try:
            fecha_obj=date.fromisoformat(fecha_iso); fecha_vto=(fecha_obj+timedelta(days=current_plazo)).isoformat()
        except: fecha_vto=fecha_iso
        exists=db.execute("SELECT id FROM facturas WHERE proveedor_id=? AND numero=?",(current_prov_id,numero)).fetchone()
        if exists: skipped+=1; continue
        try:
            db.execute("INSERT INTO facturas (proveedor_id,numero,fecha,fecha_vencimiento,importe,moneda,tipo_pago,cheque_id,categoria,estado,destino) VALUES (?,?,?,?,?,?,?,?,?,'pendiente',?)",
                (current_prov_id,numero,fecha_iso,fecha_vto,importe,current_moneda,tipo_pago,cheque_id,categoria,destino_db))
            imported+=1
        except Exception as ex: errors.append(f"{numero}: {ex}")
    db.commit(); db.close()
    return {"importadas": imported, "omitidas": skipped, "no_encontrados": not_found, "errores": errors[:10]}

@app.route('/estado_cuenta', methods=['GET','POST'])
@login_required
def estado_cuenta():
    resultado = None
    if request.method == 'POST':
        f = request.files.get('archivo')
        if not f:
            db = get_tenant_db(current_user.empresa_id)
            provs = db.execute("SELECT id, nombre FROM proveedores WHERE activo=1 ORDER BY nombre").fetchall()
            db.close()
            return render_template('estado_cuenta.html', resultado={"error": "No se subió ningún archivo."}, proveedores=provs)
        resultado = procesar_estado_cuenta(f)
    db = get_tenant_db(current_user.empresa_id)
    provs = db.execute("SELECT id, nombre FROM proveedores WHERE activo=1 ORDER BY nombre").fetchall()
    db.close()
    return render_template('estado_cuenta.html', resultado=resultado, proveedores=provs)

@app.route('/estado_cuenta/confirmar', methods=['POST'])
@login_required
def confirmar_estado_cuenta():
    proveedor_id = int(request.form.get('proveedor_id', 0))
    moneda = request.form.get('moneda', 'USD')
    facturas_raw = request.form.getlist('factura_sel')
    if not proveedor_id or not facturas_raw:
        return redirect(url_for('estado_cuenta'))
    importadas = 0; skipped = 0
    db = get_tenant_db(current_user.empresa_id)
    prov = db.execute("SELECT * FROM proveedores WHERE id=?", (proveedor_id,)).fetchone()
    plazo = prov['plazo_dias'] if prov else 30
    for raw in facturas_raw:
        try:
            parts = raw.split("|"); numero = parts[0].strip(); importe = float(parts[1])
            fecha_iso = parts[2] if len(parts) > 2 else date.today().isoformat()
            fecha_obj = date.fromisoformat(fecha_iso)
            fecha_vto = (fecha_obj + timedelta(days=plazo)).isoformat()
            numero_norm = _normalizar_numero_factura(numero)
            todas = db.execute("SELECT id, numero FROM facturas WHERE proveedor_id=?", (proveedor_id,)).fetchall()
            exists = any(_normalizar_numero_factura(r['numero']) == numero_norm for r in todas)
            if exists: skipped += 1; continue
            db.execute("INSERT INTO facturas (proveedor_id,numero,fecha,fecha_vencimiento,importe,moneda,estado) VALUES (?,?,?,?,?,?,'pendiente')",
                (proveedor_id, numero, fecha_iso, fecha_vto, importe, moneda))
            importadas += 1
        except Exception: continue
    db.commit()
    provs = db.execute("SELECT id, nombre FROM proveedores WHERE activo=1 ORDER BY nombre").fetchall()
    db.close()
    return render_template('estado_cuenta.html', resultado={"confirmado": True, "importadas": importadas, "skipped": skipped}, proveedores=provs)

def _parse_importe_uy(s):
    s=s.strip()
    if ',' in s and '.' in s: return float(s.replace('.','').replace(',','.'))
    elif ',' in s: return float(s.replace(',','.'))
    else: return float(s)

def _normalizar_numero_factura(numero):
    m = re.findall(r'\d{5,}', numero)
    return m[-1] if m else numero.strip()

def _parse_fecha_uy(s):
    from datetime import datetime as _dt
    for fmt in ('%d/%m/%Y','%d/%m/%y','%d-%m-%Y','%d-%m-%y'):
        try: return _dt.strptime(s.strip(), fmt).date().isoformat()
        except: pass
    return None

def procesar_estado_cuenta(fileobj):
    import tempfile
    if not fileobj.filename.lower().endswith(".pdf"):
        return {"error": "Solo se soportan archivos PDF."}
    try: import pdfplumber
    except ImportError: return {"error": "Instalá pdfplumber: pip install pdfplumber"}
    try:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf"); tmp_name = tmp.name; tmp.close()
        fileobj.save(tmp_name)
        texto = ""
        try:
            with pdfplumber.open(tmp_name) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t: texto += t + "\n"
        finally:
            try: os.unlink(tmp_name)
            except: pass
    except Exception as e: return {"error": f"No se pudo leer el archivo: {e}"}
    if not texto.strip(): return {"error": "No se pudo extraer texto del PDF."}
    facturas = []; seen = set()
    def agregar(numero, importe, fecha_iso, linea):
        numero = numero.strip()
        numero_norm = _normalizar_numero_factura(numero)
        if numero_norm in seen or not numero or importe is None or importe <= 0: return
        seen.add(numero_norm)
        facturas.append({"numero": numero, "importe": round(importe, 2), "fecha": fecha_iso or date.today().isoformat(), "linea_original": linea.strip()[:120]})
    lineas = texto.split("\n")
    pat_cfe = re.compile(r'(\d{2}/\d{2}/\d{2,4})\s+[^(]+\(CFE\)([A-Z]\d+)\s+U\$S\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)', re.IGNORECASE)
    for linea in lineas:
        m = pat_cfe.search(linea)
        if m:
            try: agregar(m.group(2), _parse_importe_uy(m.group(4)), _parse_fecha_uy(m.group(1)), linea)
            except: pass
    pat_unicom = re.compile(r'(\d{2}/\d{2}/\d{4})\s+(Factura|Recibo|Nota\s+de\s+[Cc]r[eé]dito)\s+([A-Z]\d{4,10})\s+(\d{2}/\d{2}/\d{4})\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)', re.IGNORECASE)
    for linea in lineas:
        m = pat_unicom.search(linea)
        if m:
            try: agregar(m.group(3), _parse_importe_uy(m.group(7)), _parse_fecha_uy(m.group(4)), linea)
            except: pass
    pat_avgus = re.compile(r'e-Factura\s+[A-Z]\s+(\d{5,10})\s+\S+\s+(\d{2}/\d{2}/\d{4})\s+([\d.,]+)\s+USD', re.IGNORECASE)
    for linea in lineas:
        m = pat_avgus.search(linea)
        if m:
            try: agregar(m.group(1), _parse_importe_uy(m.group(3)), _parse_fecha_uy(m.group(2)), linea)
            except: pass
    if not facturas:
        pat_gen = re.compile(r'(?:Factura|e-Factura|Recibo)\s+([A-Z]\d{4,10}|\d{3}/[A-Z]/\d+)', re.IGNORECASE)
        pat_imp = re.compile(r'([\d]{1,3}(?:[.,]\d{3})*[.,]\d{2})')
        pat_fec = re.compile(r'(\d{2}/\d{2}/\d{2,4})')
        for i, linea in enumerate(lineas):
            m = pat_gen.search(linea)
            if not m: continue
            numero = m.group(1); ctx = " ".join(lineas[i:i+2]); fecha_iso = None
            for fs in pat_fec.findall(ctx):
                fecha_iso = _parse_fecha_uy(fs)
                if fecha_iso: break
            importe = None
            for imp_str in pat_imp.findall(ctx):
                try:
                    v = _parse_importe_uy(imp_str)
                    if 0.5 < v < 9_999_999: importe = v; break
                except: continue
            agregar(numero, importe, fecha_iso, linea)
    return {"facturas": facturas, "texto_muestra": texto[:1000], "total_lineas": len(lineas)}

# ════════════ SALDOS A FAVOR ════════════

@app.route('/proveedores/<int:pid>/saldo-favor', methods=['GET','POST'])
@login_required
def saldo_favor_proveedor(pid):
    db = get_tenant_db(current_user.empresa_id)
    proveedor = db.execute("SELECT * FROM proveedores WHERE id=?", (pid,)).fetchone()
    if request.method == 'POST':
        f = request.form
        db.execute("""
            INSERT INTO saldos_favor (proveedor_id, fecha, monto, moneda, motivo, referencia)
            VALUES (?,?,?,?,?,?)
        """, (pid, f['fecha'], float(f['monto']), f['moneda'],
              f.get('motivo'), f.get('referencia')))
        db.commit()
        db.close()
        return redirect(url_for('proveedores'))
    saldos = db.execute("SELECT * FROM saldos_favor WHERE proveedor_id=? ORDER BY fecha DESC", (pid,)).fetchall()
    total = db.execute("SELECT COALESCE(SUM(monto),0) FROM saldos_favor WHERE proveedor_id=?", (pid,)).fetchone()[0]
    db.close()
    return render_template('saldo_favor.html', proveedor=proveedor, saldos=saldos, total=total, today=date.today().isoformat())

@app.route('/proveedores/<int:pid>/saldo-favor/<int:sid>/eliminar', methods=['POST'])
@login_required
def eliminar_saldo_favor(pid, sid):
    db = get_tenant_db(current_user.empresa_id)
    db.execute("DELETE FROM saldos_favor WHERE id=? AND proveedor_id=?", (sid, pid))
    db.commit(); db.close()
    return redirect(url_for('saldo_favor_proveedor', pid=pid))

# ════════════ BACKUP / CONFIG ════════════

@app.route('/backup')
@login_required
def backup_page():
    db = get_tenant_db(current_user.empresa_id)
    n_provs = db.execute("SELECT COUNT(*) FROM proveedores").fetchone()[0]
    n_facturas = db.execute("SELECT COUNT(*) FROM facturas").fetchone()[0]
    n_pagos = db.execute("SELECT COUNT(*) FROM pagos").fetchone()[0]
    n_cp = db.execute("SELECT COUNT(*) FROM cobros_pagos").fetchone()[0]
    db_path, _ = get_tenant_db_path(current_user.empresa_id)
    db_size = os.path.getsize(db_path) // 1024 if db_path and os.path.exists(db_path) else 0
    db.close()
    return render_template('backup.html', n_provs=n_provs, n_facturas=n_facturas,
                           n_pagos=n_pagos, n_cp=n_cp, db_size=db_size,
                           today=date.today().isoformat(), error=None)

@app.route('/backup/descargar')
@login_required
def backup_descargar():
    db_path, _ = get_tenant_db_path(current_user.empresa_id)
    if not db_path or not os.path.exists(db_path):
        abort(404)
    out = io.BytesIO()
    with open(db_path, "rb") as f: out.write(f.read())
    out.seek(0)
    return send_file(out, download_name=f"backup_{date.today().strftime('%Y%m%d')}.db", mimetype="application/octet-stream")

@app.route('/backup/restaurar', methods=['POST'])
@login_required
def backup_restaurar():
    import shutil, tempfile
    f = request.files.get('archivo')
    if not f: return redirect(url_for('backup_page'))
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".db"); f.save(tmp.name)
    try:
        conn = sqlite3.connect(tmp.name); conn.execute("SELECT COUNT(*) FROM proveedores"); conn.close()
    except Exception:
        os.unlink(tmp.name)
        db = get_tenant_db(current_user.empresa_id)
        provs = db.execute("SELECT COUNT(*) FROM proveedores").fetchone()
        db.close()
        return render_template('backup.html', error="El archivo no es una base de datos válida.",
                               n_provs=0, n_facturas=0, n_pagos=0, n_cp=0, db_size=0, today=date.today().isoformat())
    db_path, _ = get_tenant_db_path(current_user.empresa_id)
    if db_path:
        shutil.copy2(tmp.name, db_path)
    os.unlink(tmp.name)
    return redirect(url_for('index'))

@app.route('/backup/restaurar_json', methods=['POST'])
@login_required
def backup_restaurar_json():
    """Restaurar backup desde JSON exportado."""
    import tempfile
    f = request.files.get('archivo')
    if not f:
        return redirect(url_for('backup_page'))
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".json", mode='w')
    try:
        content = f.read().decode('utf-8')
        tmp.write(content)
        tmp.close()
        with open(tmp.name) as jf:
            backup = json.load(jf)
    except Exception as e:
        os.unlink(tmp.name)
        return render_template('backup.html', error=f"Error leyendo JSON: {e}",
                               n_provs=0, n_facturas=0, n_pagos=0, n_cp=0, db_size=0, today=date.today().isoformat())
    os.unlink(tmp.name)

    db = get_tenant_db(current_user.empresa_id)
    data = backup.get('data', backup)
    total = 0
    for tabla in ['saldos_favor', 'pagos', 'facturas', 'cobros_pagos', 'proveedores']:
        registros = data.get(tabla, [])
        if not registros:
            continue
        cols = list(registros[0].keys())
        placeholders = ','.join(['?'] * len(cols))
        col_names = ','.join(cols)
        for row in registros:
            vals = [row.get(c) for c in cols]
            try:
                db.execute(f"INSERT OR REPLACE INTO {tabla} ({col_names}) VALUES ({placeholders})", vals)
                total += 1
            except Exception:
                pass
    db.commit()
    db.close()

    return render_template('backup.html', success=f"Restaurados {total} registros desde JSON.",
                           n_provs=0, n_facturas=0, n_pagos=0, n_cp=0, db_size=0, today=date.today().isoformat())

@app.route('/configuracion', methods=['GET','POST'])
@login_required
def configuracion():
    cfg = load_config_tenant(current_user.empresa_id)
    saved = False; error = None
    if request.method == 'POST':
        carpeta = request.form.get('carpeta_comprobantes', '').strip()
        nombre_empresa = request.form.get('nombre_empresa', '').strip()
        if carpeta and not os.path.isdir(carpeta):
            error = f"La carpeta '{carpeta}' no existe."
        else:
            save_config_tenant(current_user.empresa_id, {"carpeta_comprobantes": carpeta, "nombre_empresa": nombre_empresa})
            cfg = load_config_tenant(current_user.empresa_id)
            saved = True
    return render_template('configuracion.html', cfg=cfg, saved=saved, error=error)

@app.route('/api/buscar_comprobante')
@login_required
def api_buscar_comprobante():
    prov = request.args.get('proveedor', '')
    fecha = request.args.get('fecha', '')
    archivos = buscar_comprobante(prov, fecha, current_user.empresa_id)
    return jsonify([{"ruta": r, "nombre": n} for r, n in archivos])

# ════════════ MAIN ════════════

# ════════════ BACKUP TENANT (admin restaura backup de cualquier empresa) ════════════

@app.route('/admin/empresa/<int:eid>/restaurar-backup', methods=['POST'])
@login_required
@admin_required
def admin_restaurar_backup_tenant(eid):
    import shutil, tempfile
    f = request.files.get('archivo')
    if not f:
        flash('No se seleccionó ningún archivo.', 'danger')
        return redirect(url_for('admin_panel'))

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
    f.save(tmp.name)
    try:
        conn = sqlite3.connect(tmp.name)
        conn.execute('SELECT COUNT(*) FROM proveedores')
        conn.close()
    except Exception:
        os.unlink(tmp.name)
        flash('El archivo no es una base de datos válida.', 'danger')
        return redirect(url_for('admin_panel'))

    db_path, slug = get_tenant_db_path(eid)
    if db_path:
        # Backup del backup actual antes de restaurar
        if os.path.exists(db_path):
            backup_prev = db_path + '.prev'
            shutil.copy2(db_path, backup_prev)
        shutil.copy2(tmp.name, db_path)
        os.unlink(tmp.name)
        flash(f'Backup restaurado en "{slug}". Si algo salió mal, el backup anterior se guardó como .prev', 'success')
    else:
        flash('Empresa no encontrada.', 'danger')
    return redirect(url_for('admin_panel'))

@app.route('/admin/empresa/<int:eid>/descargar-backup')
@login_required
@admin_required
def admin_descargar_backup_tenant(eid):
    db_path, slug = get_tenant_db_path(eid)
    if not db_path or not os.path.exists(db_path):
        abort(404)
    out = io.BytesIO()
    with open(db_path, 'rb') as f:
        out.write(f.read())
    out.seek(0)
    return send_file(out, download_name=f'backup_{slug}_{date.today().strftime("%Y%m%d")}.db',
                     mimetype='application/octet-stream')

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', '0') == '1'
    print(f"\n✅ Sistema Unificado SaaS v3.0 → http://localhost:{port}\n")
    app.run(debug=debug, host='0.0.0.0', port=port)
